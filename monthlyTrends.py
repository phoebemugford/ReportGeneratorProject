import os
import pandas as pd
import requests
import json
from datetime import datetime
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util import Retry
from concurrent.futures import ThreadPoolExecutor, as_completed
import traceback
import tkinter as tk
from tkinter import filedialog, ttk, messagebox


def monthly_trends_report(save_location_entry, year_chosen_entry, work_hours_only_var, spaces_to_include_entry,
                          spaces_df, window,
                          report_track):
    AUTH = "Authentication Key"
    URL = "API.com"
    HEADERS = {"Authorization": f"Bearer {AUTH}"}
    SESSION = requests.Session()
    RETRY = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
    SESSION.mount('http://', HTTPAdapter(max_retries=RETRY))
    SESSION.mount('https://', HTTPAdapter(max_retries=RETRY))

    # Gets data for each device
    def get_devices(space_id, year, work_hours_only):
        url = f"{URL}devices&space.id={space_id}&limit=2000"
        res = SESSION.get(url, headers=HEADERS)
        print(res.status_code)
        data_json = json.loads(res.content)
        devices = data_json["devices"]
        time.sleep(1)
        devices_info = []
        for device in devices:
            product_model = device.get('product', {}).get('model') if device.get('product') else None
            # Sets general details and defaults for each device's info
            device_info = {'id': device.get('id'), 'serialNumber': device.get('serialNumber'),
                           'name': device.get('name'),
                           'isOnline': device.get('isOnline'), 'product_model': product_model, 'space_id': space_id,
                           'space_name': device.get('space', {}).get('name'), '1CO2': "N/A", '1Temp': "N/A",
                           '1Hum': "N/A",
                           '1ACH': "N/A", '1VP_Diff': "N/A", '2CO2': "N/A", '2Temp': "N/A", '2Hum': "N/A",
                           '2ACH': "N/A",
                           '2VP_Diff': "N/A", '3CO2': "N/A", '3Temp': "N/A", '3Hum': "N/A", '3ACH': "N/A",
                           '3VP_Diff': "N/A", '4CO2': "N/A", '4Temp': "N/A", '4Hum': "N/A", '4ACH': "N/A",
                           '4VP_Diff': "N/A", '5CO2': "N/A", '5Temp': "N/A", '5Hum': "N/A", '5ACH': "N/A",
                           '5VP_Diff': "N/A", '6CO2': "N/A", '6Temp': "N/A", '6Hum': "N/A", '6ACH': "N/A",
                           '6VP_Diff': "N/A", '7CO2': "N/A", '7Temp': "N/A", '7Hum': "N/A", '7ACH': "N/A",
                           '7VP_Diff': "N/A", '8CO2': "N/A", '8Temp': "N/A", '8Hum': "N/A", '8ACH': "N/A",
                           '8VP_Diff': "N/A", '9CO2': "N/A", '9Temp': "N/A", '9Hum': "N/A", '9ACH': "N/A",
                           '9VP_Diff': "N/A", '10CO2': "N/A", '10Temp': "N/A", '10Hum': "N/A", '10ACH': "N/A",
                           '10VP_Diff': "N/A", '11CO2': "N/A", '11Temp': "N/A", '11Hum': "N/A", '11ACH': "N/A",
                           '11VP_Diff': "N/A", '12CO2': "N/A", '12Temp': "N/A", '12Hum': "N/A", '12ACH': "N/A",
                           '12VP_Diff': "N/A"}
            device_info[
                'Link to Device'] = f"dashboard.com/{device_info['space_id']}/devices/{device_info['id']}"
            if product_model:
                if work_hours_only == 1:
                    open_hours = "true"
                else:
                    open_hours = "false"
                if product_model.startswith('AMB_'):  # Only processes devices whose product model starts with AMB_
                    # Sets month to current month if current year, otherwise December
                    current_year = int(datetime.today().strftime('%Y'))
                    if current_year == year:
                        month = int(datetime.today().strftime('%m'))
                    else:
                        month = 12
                    sensor = 'co2'
                    while month >= 1:  # Loops through all months since (and including) January
                        isDone = False
                        if month == 12:  # If in December, next month is January
                            next_month = 1
                            end_year = year + 1
                        else:
                            next_month = month + 1
                            end_year = year
                        event_url = f'{URL}event/device.id={device_info["id"]}&name={sensor}&processed=true' \
                                    f'&resolutionFrequency=1&onlySpaceOpeningHours={open_hours}&since={year}-{month}-01&until' \
                                    f'={end_year}-{next_month}-01&limit=1'  # Gets data to get event count
                        event_res = SESSION.get(event_url, headers=HEADERS)
                        print(res.status_code)
                        event_data_json = json.loads(event_res.content)
                        event_count = event_data_json['count']
                        if event_count > 0:  # If there is data
                            while not isDone:
                                event_url = f'{URL}event/device.id={device_info["id"]}&name={sensor}&processed=true' \
                                            f'&resolutionFrequency=1&onlySpaceOpeningHours={open_hours}&since={year}-{month}' \
                                            f'-01&until={end_year}-{next_month}-01&limit={event_count}'
                                event_res = SESSION.get(event_url, headers=HEADERS)
                                event_data_json = json.loads(event_res.content)
                                updated_event_count = event_data_json['count']
                                if updated_event_count > event_count and updated_event_count > 0:
                                    # Gets data again if there is more data for a certain parameter
                                    event_url = f'{URL}event/device.id={device_info["id"]}&name={sensor}&processed=true' \
                                                f'&resolutionFrequency=1&onlySpaceOpeningHours={open_hours}&since={year}-{month}' \
                                                f'-01&until={end_year}-{next_month}-01&limit={updated_event_count}'
                                    event_res = SESSION.get(event_url, headers=HEADERS)
                                    event_data_json = json.loads(event_res.content)
                                if 'result' in event_data_json and event_data_json['result'] and \
                                        updated_event_count > 0:
                                    # Gets average using nested list
                                    df_nested_list = pd.json_normalize(event_data_json, record_path=['result'])
                                    average_reading = df_nested_list['data'].mean()
                                    # Assigns data to device_info
                                    if sensor == 'co2':
                                        average_reading = round(average_reading, 0)
                                        device_info[f"{month}CO2"] = average_reading
                                    elif sensor == 'humidity':
                                        average_reading = round(average_reading, 1)
                                        device_info[f"{month}Hum"] = average_reading
                                    elif sensor == 'ACH_runav7':
                                        average_reading = round(average_reading, 1)
                                        device_info[f"{month}ACH"] = average_reading
                                    elif sensor == 'VP_diff_runav7':
                                        average_reading = round(average_reading, 1)
                                        device_info[f"{month}VP_Diff"] = average_reading
                                    elif sensor == 'temperature':
                                        average_reading = round(average_reading, 1)
                                        device_info[f"{month}Temp"] = average_reading
                                    else:
                                        isDone = True

                                # Moves onto next parameter or ends loop
                                if sensor == 'co2':
                                    sensor = 'humidity'
                                elif sensor == 'humidity':
                                    sensor = 'ACH_runav7'
                                elif sensor == 'ACH_runav7':
                                    sensor = 'VP_diff_runav7'
                                elif sensor == 'VP_diff_runav7':
                                    sensor = 'temperature'
                                elif sensor == 'temperature':
                                    isDone = True
                                    sensor = 'co2'
                                else:
                                    isDone = True
                        month = month - 1  # Goes to next month
                    devices_info.append(device_info)
        devices_df = pd.DataFrame(devices_info)
        return devices_df

    def apply_header_style(ws):
        # Define style for header
        header = NamedStyle(name="header")
        header.font = Font(bold=True)
        header.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                               top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        header.alignment = Alignment(horizontal="center")

        # Apply style to headers
        for cell in ws[1]:
            cell.style = header

    # Function to create the Excel file
    def create_excel_file(df, save_location, year_chosen, work_hours_only, spaces_to_include):
        # Sets variables for filename
        if len(year_chosen) == 0:
            year = datetime.today().strftime('%Y')
        else:
            year = year_chosen
        if work_hours_only == 1:
            work_all = "work"
        else:
            work_all = "all"
        if len(spaces_to_include) == 0:
            filename = f"devices-monthly-info-{year}-{work_all}-hours.xlsx"
        else:
            filename = f"devices-monthly-info-{year}-{work_all}-hours-{spaces_to_include}.xlsx"
        save_path = os.path.join(save_location, filename)
        df.to_excel(save_path, index=False)
        return save_path

    # Function to generate and format the report
    def generate_monthly_report(save_location, year_chosen, work_hours_only, spaces_to_include, spaces):
        if len(year_chosen) == 0:
            year = int(datetime.today().strftime('%Y'))
        else:
            year = int(year_chosen)
        try:
            all_devices_df = pd.DataFrame()

            num_spaces = len(spaces_df["id"])
            progress_bar["maximum"] = num_spaces

            with ThreadPoolExecutor(max_workers=2) as executor:
                future_to_space = {executor.submit(get_devices, space_id, year, work_hours_only): space_id for space_id
                                   in
                                   spaces["id"]}

                for future in as_completed(future_to_space):
                    space_id = future_to_space[future]
                    try:
                        devices_df = future.result()
                    except Exception as e:
                        print(f"Error occurred while getting devices for '{space_id}': {str(e)}")
                        traceback.print_exc()
                        continue

                    if not devices_df.empty:  # Only append if devices_df is not empty
                        all_devices_df = pd.concat([all_devices_df, devices_df], ignore_index=True)

                    progress_bar.step(1)
                    window.update()

            # Then reorder the DataFrame columns
            all_devices_df = all_devices_df[
                [
                    'space_name',
                    'name',
                    'Link to Device',
                    '1CO2',
                    '1Temp',
                    '1Hum',
                    '1ACH',
                    '1VP_Diff',
                    '2CO2',
                    '2Temp',
                    '2Hum',
                    '2ACH',
                    '2VP_Diff',
                    '3CO2',
                    '3Temp',
                    '3Hum',
                    '3ACH',
                    '3VP_Diff',
                    '4CO2',
                    '4Temp',
                    '4Hum',
                    '4ACH',
                    '4VP_Diff',
                    '5CO2',
                    '5Temp',
                    '5Hum',
                    '5ACH',
                    '5VP_Diff',
                    '6CO2',
                    '6Temp',
                    '6Hum',
                    '6ACH',
                    '6VP_Diff',
                    '7CO2',
                    '7Temp',
                    '7Hum',
                    '7ACH',
                    '7VP_Diff',
                    '8CO2',
                    '8Temp',
                    '8Hum',
                    '8ACH',
                    '8VP_Diff',
                    '9CO2',
                    '9Temp',
                    '9Hum',
                    '9ACH',
                    '9VP_Diff',
                    '10CO2',
                    '10Temp',
                    '10Hum',
                    '10ACH',
                    '10VP_Diff',
                    '11CO2',
                    '11Temp',
                    '11Hum',
                    '11ACH',
                    '11VP_Diff',
                    '12CO2',
                    '12Temp',
                    '12Hum',
                    '12ACH',
                    '12VP_Diff'
                ]
            ]

            # Sort DataFrame by 'space_name'
            all_devices_df.sort_values('space_name', ascending=True, inplace=True)

            # Reset index after sorting
            all_devices_df.reset_index(drop=True, inplace=True)

            save_path = create_excel_file(all_devices_df, save_location, year_chosen, work_hours_only,
                                          spaces_to_include)

            # Opens sheet to be edited
            wb = load_workbook(save_path)
            ws = wb['Sheet1']
            ws.insert_rows(idx=1)
            apply_header_style(ws)

            # Finds number of columns and adds link to device space to device name
            name_column_found = False
            link_column_found = False
            column_counter = 1
            empty_checker = ws.cell(row=2, column=1).value
            while empty_checker is not None or not link_column_found or not name_column_found:
                empty_checker = ws.cell(row=2, column=column_counter).value
                column_letter = get_column_letter(column_counter)
                if not name_column_found and empty_checker.startswith("name"):
                    name_column_found = True
                    name_column_letter = column_letter
                if not link_column_found and "Link" in empty_checker:
                    link_column_found = True
                    link_column_letter = column_letter
                    link_column_number = column_counter
                column_counter = column_counter + 1
            column_counter = column_counter - 1
            rows_devs = len(all_devices_df["name"]) + 2
            for row in range(3, rows_devs + 1):
                link = ws[f'{link_column_letter}{row}'].value
                name = ws[f'{name_column_letter}{row}'].value
                ws[f'{name_column_letter}{row}'].hyperlink = link
                ws[f'{name_column_letter}{row}'].value = name
                ws[f'{name_column_letter}{row}'].style = "Hyperlink"
            ws.delete_cols(link_column_number)
            column_counter = column_counter - 1
            rows_devs = len(all_devices_df["name"]) + 2

            # Creates borders, fonts, and backgrounds
            thin_border = Border(left=Side(style='thin', color="D9D9D9D9"),
                                 right=Side(style='thin', color="D9D9D9D9"),
                                 top=Side(style='thin', color="D9D9D9D9"),
                                 bottom=Side(style='thin', color="D9D9D9D9"))
            no_top_thin_border = Border(left=Side(style='thin', color="D9D9D9D9"),
                                        right=Side(style='thin', color="D9D9D9D9"),
                                        bottom=Side(style='thin', color="D9D9D9D9"))
            rest_gray_left_black_border = Border(left=Side(style='thin', color="FF000000"),
                                                 right=Side(style='thin', color="D9D9D9D9"),
                                                 bottom=Side(style='thin', color="D9D9D9D9"))
            left_black_thin_border = Border(left=Side(style='thin', color="FF000000"))
            black_font = Font(color="FF000000", italic=False)
            black_italic_font = Font(color="FF000000", italic=True)
            white_font = Font(color="FFFFFFFF", italic=False)
            white_background = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type='solid')
            black_background = PatternFill(start_color="FF000000", end_color="FF000000", fill_type='solid')
            red_background = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type='solid')
            yellow_background = PatternFill(start_color="00FFC000", end_color="00FFC000", fill_type='solid')
            blue_background = PatternFill(start_color="0000B0F0", end_color="0000B0F0", fill_type='solid')
            purple_background = PatternFill(start_color="00D55EFF", end_color="00D55EFF")

            # Stops N/A cells from being colored and applies correct border
            for col in range(1, column_counter):
                column_letter = get_column_letter(col)
                if "CO2" in ws.cell(row=2, column=col).value:  # CO2 cells get left border to differentiate months
                    if rows_devs > 3:
                        NA_style = DifferentialStyle(fill=white_background, border=left_black_thin_border,
                                                     font=black_font)
                        rule = Rule(type='containsText', operator='containsText', text='/', dxf=NA_style)
                        rule.formula = [f'NOT(ISERROR(SEARCH("/",{column_letter}4)))']
                        ws.conditional_formatting.add(f"{column_letter}4:{column_letter}{rows_devs}", rule)
                    NA_style = DifferentialStyle(fill=white_background, border=rest_gray_left_black_border,
                                                 font=black_font)
                    rule = Rule(type='containsText', operator='containsText', text='/', dxf=NA_style)
                    rule.formula = [f'NOT(ISERROR(SEARCH("/",{column_letter}3)))']
                    ws.conditional_formatting.add(f"{column_letter}3", rule)
                if rows_devs > 3:  # If more than one device
                    NA_style = DifferentialStyle(fill=white_background, border=thin_border, font=black_font)
                    rule = Rule(type='containsText', operator='containsText', text='/', dxf=NA_style)
                    rule.formula = [f'NOT(ISERROR(SEARCH("/",{column_letter}4)))']
                    ws.conditional_formatting.add(f"{column_letter}4:{column_letter}{rows_devs}", rule)
                # Top row of data does not have top border to keep the headers' border intact
                NA_style = DifferentialStyle(fill=white_background, border=no_top_thin_border, font=black_font)
                rule = Rule(type='containsText', operator='containsText', text='/', dxf=NA_style)
                rule.formula = [f'NOT(ISERROR(SEARCH("/",{column_letter}3)))']
                ws.conditional_formatting.add(f"{column_letter}3:{column_letter}3", rule)

            # Colors cells and renames column headers
            for col in range(1, column_counter):
                column_letter = get_column_letter(col)
                # Resets styles
                out_of_range_style = DifferentialStyle(fill=purple_background, font=black_italic_font)
                black_style = DifferentialStyle(fill=black_background, font=white_font)
                red_style = DifferentialStyle(fill=red_background, font=black_font)
                yellow_style = DifferentialStyle(fill=yellow_background, font=black_font)
                blue_style = DifferentialStyle(fill=blue_background, font=black_font)

                current_cell = ws.cell(row=2, column=col).value
                if "CO2" in current_cell:  # Renames and colors CO2 columns
                    merge_start = column_letter
                    ws[f'{column_letter}2'] = "CO2"
                    out_of_range_style = DifferentialStyle(fill=purple_background, font=black_italic_font,
                                                           border=left_black_thin_border)
                    rule = Rule(type='cellIs', operator='greaterThan', formula=[10000], dxf=out_of_range_style)
                    rule.formula = [f"{column_letter}3>10000"]  # CO2 > 10000
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    black_style = DifferentialStyle(fill=black_background, font=white_font,
                                                    border=left_black_thin_border)
                    rule = Rule(type="expression", dxf=black_style)
                    rule.formula = [f"{column_letter}3>1500"]  # 10001 > CO2 > 1500
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    red_style = DifferentialStyle(fill=red_background, font=black_font, border=left_black_thin_border)
                    rule = Rule(type="expression", dxf=red_style)
                    rule.formula = [f"{column_letter}3>1000"]  # 1501 > CO2 > 1000
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    yellow_style = DifferentialStyle(fill=yellow_background, font=black_font,
                                                     border=left_black_thin_border)
                    rule = Rule(type="expression", dxf=yellow_style)
                    rule.formula = [f"{column_letter}3>800"]  # 1001 > CO2 > 800
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    border_style = DifferentialStyle(border=left_black_thin_border)
                    rule = Rule(type="expression", dxf=border_style)
                    rule.formula = [f"{column_letter}3>=400"]  # Adds left border to rest of CO2 cells
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=out_of_range_style)
                    rule.formula = [f"{column_letter}3<400"]  # Adds left border to rest of CO2 cells
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                elif "Temp" in current_cell:  # Renames and colors temperature columns
                    ws[f'{column_letter}2'] = "Temp"
                    rule = Rule(type='cellIs', operator='greaterThan', formula=[40], dxf=out_of_range_style)
                    rule.formula = [f"{column_letter}3>40"]  # Temp > 40
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=red_style)
                    rule.formula = [f"{column_letter}3>25"]  # 41 > Temp > 25
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=blue_style)
                    rule.formula = [f"{column_letter}3<16"]  # 16 > Temp > -21
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=out_of_range_style)
                    rule.formula = [f"{column_letter}3<-20"]  # -20 > Temp
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                elif "Hum" in current_cell:  # Renames and colors humidity columns
                    ws[f'{column_letter}2'] = "Humidity"
                    rule = Rule(type='cellIs', operator='greaterThan', formula=[99], dxf=out_of_range_style)
                    rule.formula = [f"{column_letter}3>99"]  # Humidity > 99
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=red_style)
                    rule.formula = [f"{column_letter}3>60"]  # 100 > Humidity > 60
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=yellow_style)
                    rule.formula = [f"{column_letter}3<40"]  # 40 > Humidity > 4
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=out_of_range_style)
                    rule.formula = [f"{column_letter}3<5"]  # 5 > Humidity
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                elif "ACH" in current_cell:  # Renames ACH columns
                    ws[f'{column_letter}2'] = "ACH"
                elif "VP_Diff" in current_cell:  # Renames and colors VP Diff columns and merges/names month name cells
                    merge_end = column_letter
                    ws.merge_cells(f'{merge_start}1:{merge_end}1')
                    # Names month header cells
                    if "12" in current_cell:
                        ws[f"{merge_start}1"] = "December"
                    elif "11" in current_cell:
                        ws[f"{merge_start}1"] = "November"
                    elif "10" in current_cell:
                        ws[f"{merge_start}1"] = "October"
                    elif "9" in current_cell:
                        ws[f"{merge_start}1"] = "September"
                    elif "8" in current_cell:
                        ws[f"{merge_start}1"] = "August"
                    elif "7" in current_cell:
                        ws[f"{merge_start}1"] = "July"
                    elif "6" in current_cell:
                        ws[f"{merge_start}1"] = "June"
                    elif "5" in current_cell:
                        ws[f"{merge_start}1"] = "May"
                    elif "4" in current_cell:
                        ws[f"{merge_start}1"] = "April"
                    elif "3" in current_cell:
                        ws[f"{merge_start}1"] = "March"
                    elif "2" in current_cell:
                        ws[f"{merge_start}1"] = "February"
                    elif "1" in current_cell:
                        ws[f"{merge_start}1"] = "January"
                    ws[f'{column_letter}2'] = "VP Diff"
                    rule = Rule(type="expression", dxf=black_style)
                    rule.formula = [f"{column_letter}3>0.8"]  # VP Diff > 0.8
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=red_style)
                    rule.formula = [f"{column_letter}3>0.6"]  # 0.8 >= VP Diff > 0.6
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=yellow_style)
                    rule.formula = [f"{column_letter}3>0.3"]  # 0.6 >= VP Diff > 0.3
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)

            ws.freeze_panes = 'C3'  # Freezes top two rows and left two columns
            wb.save(save_path)  # Saves and closes spreadsheet

            print(f"Report generated successfully: {save_path}")

        except Exception as e:
            print(f"An error occurred: {str(e)}")
            traceback.print_exc()

    # "Generating report x/x" label
    report_track_label = tk.Label(window, text=report_track)
    report_track_label.grid(column=0, row=23, pady=3)

    # Progress bar
    progress_bar = ttk.Progressbar(window, orient=tk.HORIZONTAL, length=300, mode='determinate')
    progress_bar.grid(column=0, row=24, pady=3)

    generate_monthly_report(save_location_entry, year_chosen_entry, work_hours_only_var, spaces_to_include_entry,
                            spaces_df)

    # Removes label and progress bar
    report_track_label.grid_remove()
    progress_bar.grid_remove()
