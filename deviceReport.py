import os
import pandas as pd
import requests
import json
from datetime import datetime
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util import Retry
from concurrent.futures import ThreadPoolExecutor, as_completed
import traceback
import tkinter as tk
from tkinter import filedialog, ttk, messagebox


def device_overview_report(save_location_entry, start_date_entry, end_date_entry, work_hours_only_var,
                           spaces_to_include_entry, spaces_df, window, report_track):
    AUTH = "Authentication Key"
    URL = "API.com"
    HEADERS = {"Authorization": f"Bearer {AUTH}"}
    SESSION = requests.Session()
    RETRY = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
    SESSION.mount('http://', HTTPAdapter(max_retries=RETRY))
    SESSION.mount('https://', HTTPAdapter(max_retries=RETRY))

    # Gets data for each device
    def get_devices(space_id, start_date_mill, end_date_mill, work_hours_only):
        url = f"{URL}devices&space.id={space_id}&limit=2000"
        res = SESSION.get(url, headers=HEADERS)
        print(res.status_code)
        data_json = json.loads(res.content)
        devices = data_json["devices"]
        time.sleep(1)
        devices_info = []
        for device in devices:
            # Sets general details and defaults for each device's info
            product_model = device.get('product', {}).get('model') if device.get('product') else None
            device_info = {'id': device.get('id'), 'serialNumber': device.get('serialNumber'),
                           'name': device.get('name'),
                           'isOnline': device.get('isOnline'), 'product_model': product_model, 'space_id': space_id,
                           'space_name': device.get('space', {}).get('name'), 'sensorT': "N/A", 'aveTemp': "N/A",
                           'maxReadingTemp': "N/A",
                           'dateOfMaxTemp': "N/A", 'timeOfMaxTemp': "N/A", 'minReadingTemp': "N/A",
                           'dateOfMinTemp': "N/A",
                           'timeOfMinTemp': "N/A", 'sensorH': "N/A", 'aveHumidity': "N/A", 'maxReadingHumidity': "N/A",
                           'dateOfMaxHumidity': "N/A", 'timeOfMaxHumidity': "N/A", 'minReadingHumidity': "N/A",
                           'dateOfMinHumidity': "N/A", 'timeOfMinHumidity': "N/A", 'sensorC': "N/A", 'aveCO2': "N/A",
                           'maxReadingCO2': "N/A",
                           'dateOfMaxCO2': "N/A", 'timeOfMaxCO2': "N/A", 'minReadingCO2': "N/A", 'dateOfMinCO2': "N/A",
                           'timeOfMinCO2': "N/A", '%timeInHighRiskCO2': "N/A", '%timeInHighRiskHumidity': "N/A",
                           '%timeInHighRiskTemp': "N/A", '%timeInHighRiskVP_Diff': "N/A", 'sensorACH': "N/A",
                           'aveACH': "N/A", 'maxReadingACH': "N/A", 'dateOfMaxACH': "N/A", 'timeOfMaxACH': "N/A",
                           'minReadingACH': "N/A",
                           'dateOfMinACH': "N/A", 'timeOfMinACH': "N/A", 'sensorVP': "N/A", 'aveVP_Diff': "N/A",
                           'maxReadingVP_Diff': "N/A", 'dateOfMaxVP_Diff': "N/A", 'timeOfMaxVP_Diff': "N/A",
                           'minReadingVP_Diff': "N/A",
                           'dateOfMinVP_Diff': "N/A", 'timeOfMinVP_Diff': "N/A", 'Fuel Poverty Risk': "N/A",
                           'Void Risk': "N/A"}
            device_info[
                'Link to Device'] = f"dashboard.com/{device_info['space_id']}/devices/{device_info['id']}"
            if product_model:
                if work_hours_only == 1:
                    open_hours = "true"
                else:
                    open_hours = "false"
                if product_model.startswith('AMB_'):  # Only processes devices whose product model starts with AMB_
                    sensor = 'co2'
                    event_url = f'{URL}event/device.id={device_info["id"]}&name={sensor}&processed=true' \
                                f'&resolutionFrequency=1&onlySpaceOpeningHours={open_hours}&since={start_date_mill}&until={end_date_mill}&limit=1'
                    event_res = SESSION.get(event_url, headers=HEADERS)
                    print(res.status_code)
                    event_data_json = json.loads(event_res.content)
                    event_count = event_data_json['count']
                    if event_count > 0:  # If there is data
                        isDone = False
                        while not isDone:
                            # Gets data for co2, temp, and humidity
                            event_url = f'{URL}event/device.id={device_info["id"]}&name={sensor}&processed=true' \
                                        f'&resolutionFrequency=1&onlySpaceOpeningHours={open_hours}&since={start_date_mill}&until={end_date_mill}&limit={event_count}'
                            event_res = SESSION.get(event_url, headers=HEADERS)
                            print(res.status_code)
                            event_data_json = json.loads(event_res.content)
                            updated_event_count = event_data_json['count']
                            if updated_event_count > event_count and updated_event_count > 0:
                                # Gets data again if greater count
                                event_url = f'{URL}event/device.id={device_info["id"]}&name={sensor}&processed=true' \
                                            f'&resolutionFrequency=1&onlySpaceOpeningHours={open_hours}&since={start_date_mill}' \
                                            f'&until={end_date_mill}&limit={updated_event_count}'
                                event_res = SESSION.get(event_url, headers=HEADERS)
                                event_data_json = json.loads(event_res.content)
                            if 'result' in event_data_json and event_data_json['result'] and updated_event_count > 0:
                                # Gets ave, min, and max using nested list
                                df_nested_list = pd.json_normalize(event_data_json, record_path=['result'])
                                df_nested_list['data'] = pd.to_numeric(df_nested_list['data'])
                                average_reading = df_nested_list['data'].mean()
                                max_reading = df_nested_list['data'].max()
                                min_reading = df_nested_list['data'].min()
                                max_reading_index = df_nested_list['data'].idxmax()
                                min_reading_index = df_nested_list['data'].idxmin()

                                # Gets number of data points above certain threshold (by sensor)
                                risk_number = 0
                                if sensor == 'co2':
                                    limit = 1500
                                    risk_number = df_nested_list['data'][df_nested_list['data'] > limit].count()
                                elif sensor == 'humidity':
                                    limit = 60
                                    risk_number = df_nested_list['data'][df_nested_list['data'] > limit].count()
                                elif sensor == 'temperature':
                                    limit = 16
                                    risk_number = df_nested_list['data'][df_nested_list['data'] < limit].count()
                                elif sensor == 'VP_diff_runav7':
                                    limit = 0.8
                                    risk_number = df_nested_list['data'][df_nested_list['data'] > limit].count()
                                risk_number = 100 * risk_number / updated_event_count
                                device_info['sensor'] = sensor
                                risk_number = round(risk_number, 2)

                                # Gets date and time of max/min values from index
                                if isinstance(max_reading_index, int):
                                    if isinstance(event_data_json['result'][max_reading_index].get('timestamp'), str):
                                        timestamp = datetime.fromisoformat(
                                            event_data_json['result'][max_reading_index].get('timestamp'))
                                        max_date = " " + timestamp.strftime('%d/%m/%Y')
                                        max_time = " " + timestamp.strftime('%H:%M')
                                else:
                                    max_date = "N/A"
                                    max_time = "N/A"
                                if isinstance(min_reading_index, int):
                                    if isinstance(event_data_json['result'][min_reading_index].get('timestamp'), str):
                                        timestamp = datetime.fromisoformat(
                                            event_data_json['result'][min_reading_index].get('timestamp'))
                                        min_date = " " + timestamp.strftime('%d/%m/%Y')
                                        min_time = " " + timestamp.strftime('%H:%M')
                                else:
                                    min_date = "N/A"
                                    min_time = "N/A"

                                # Assigns data to device_info
                                if sensor == 'co2':
                                    device_info['sensorC'] = 'co2'
                                    average_reading = round(average_reading, 0)
                                    device_info['aveCO2'] = average_reading
                                    device_info['maxReadingCO2'] = max_reading
                                    device_info['dateOfMaxCO2'] = max_date
                                    device_info['timeOfMaxCO2'] = max_time
                                    device_info['minReadingCO2'] = min_reading
                                    device_info['dateOfMinCO2'] = min_date
                                    device_info['timeOfMinCO2'] = min_time
                                    device_info['%timeInHighRiskCO2'] = f"{risk_number}%"
                                elif sensor == 'humidity':
                                    device_info['sensorH'] = 'humidity'
                                    average_reading = round(average_reading, 1)
                                    device_info['aveHumidity'] = average_reading
                                    device_info['maxReadingHumidity'] = max_reading
                                    device_info['dateOfMaxHumidity'] = max_date
                                    device_info['timeOfMaxHumidity'] = max_time
                                    device_info['minReadingHumidity'] = min_reading
                                    device_info['dateOfMinHumidity'] = min_date
                                    device_info['timeOfMinHumidity'] = min_time
                                    device_info['%timeInHighRiskHumidity'] = f"{risk_number}%"
                                elif sensor == 'ACH_runav7':
                                    device_info['sensorACH'] = "ACH"
                                    average_reading = round(average_reading, 1)
                                    device_info["aveACH"] = average_reading
                                    device_info['maxReadingACH'] = max_reading
                                    device_info['dateOfMaxACH'] = max_date
                                    device_info['timeOfMaxACH'] = max_time
                                    device_info['minReadingACH'] = min_reading
                                    device_info['dateOfMinACH'] = min_date
                                    device_info['timeOfMinACH'] = min_time
                                elif sensor == 'VP_diff_runav7':
                                    device_info['sensorVP'] = "VP diff"
                                    average_reading = round(average_reading, 1)
                                    device_info["aveVP_Diff"] = average_reading
                                    device_info['maxReadingVP_Diff'] = max_reading
                                    device_info['dateOfMaxVP_Diff'] = max_date
                                    device_info['timeOfMaxVP_Diff'] = max_time
                                    device_info['minReadingVP_Diff'] = min_reading
                                    device_info['dateOfMinVP_Diff'] = min_date
                                    device_info['timeOfMinVP_Diff'] = min_time
                                    device_info['%timeInHighRiskVP_Diff'] = f"{risk_number}%"
                                elif sensor == 'temperature':
                                    device_info['sensorT'] = 'temp'
                                    average_reading = round(average_reading, 1)
                                    device_info['aveTemp'] = average_reading
                                    device_info['maxReadingTemp'] = max_reading
                                    device_info['dateOfMaxTemp'] = max_date
                                    device_info['timeOfMaxTemp'] = max_time
                                    device_info['minReadingTemp'] = min_reading
                                    device_info['dateOfMinTemp'] = min_date
                                    device_info['timeOfMinTemp'] = min_time
                                    device_info['%timeInHighRiskTemp'] = f"{risk_number}%"
                                else:
                                    isDone = True

                            # Moves onto next parameter or ends loop
                            if sensor == 'co2':
                                sensor = 'humidity'
                            elif sensor == 'humidity':
                                sensor = 'ACH_runav7'
                            elif sensor == 'ACH_runav7':
                                sensor = 'VP_diff_runav7'
                            elif sensor == 'VP_diff_runav7':
                                sensor = 'temperature'
                            elif sensor == 'temperature':
                                isDone = True
                                sensor = 'co2'
                            else:
                                isDone = True
                    else:
                        print("No data for device. Moving to next.")

                    # Gets fuel poverty risk and void risk information
                    fuel_poverty = get_events(device_info['id'], "flag_fuel_poverty_risk", start_date_mill,
                                              end_date_mill)
                    if fuel_poverty:
                        device_info['Fuel Poverty Risk'] = fuel_poverty
                    void_risk = get_events(device_info['id'], "flag_void_risk", start_date_mill, end_date_mill)
                    if void_risk:
                        device_info['Void Risk'] = void_risk

                    devices_info.append(device_info)

        devices_df = pd.DataFrame(devices_info)
        return devices_df

    # Gets most recent data point from unprocessed event
    def get_events(device_id, event, start_date, end_date):
        base_url = "API.com/event"
        params = {
            "device.id": device_id,
            "name": event,
            "unprocessed": "true",
            "resolutionFrequency": "1",
            "since": start_date,
            "until": end_date,
            "limit": "20",
        }

        response = requests.get(base_url, params=params, headers=HEADERS)

        if response.status_code == 200:
            data = response.json()
            if data['count'] > 0:
                return data['events'][0]['data']
            else:
                return None
        else:
            print(f"Failed to fetch data. Status code: {response.status_code}")
            return None

    # Function to apply header style to file
    def apply_header_style(ws):
        # Define style for header
        header = NamedStyle(name="header")
        header.font = Font(bold=True, italic=False)
        header.border = Border(left=Side(border_style="medium"), right=Side(border_style="medium"),
                               bottom=Side(border_style="thin"))
        header.alignment = Alignment(horizontal="center")

        # Apply style to headers
        for cell in ws[1]:
            cell.style = header

    # Function to create the Excel file
    def create_excel_file(df, save_location, start_date, end_date, work_hours_only, spaces_to_include):
        # Sets variables for filename
        start_date_filename = datetime.strftime(start_date, '%d-%m-%Y')
        end_date_filename = datetime.strftime(end_date, '%d-%m-%Y')
        if work_hours_only == 1:
            work_all = "work"
        else:
            work_all = "all"
        if len(spaces_to_include) == 0:
            filename = f"devices-info-{start_date_filename}-to-{end_date_filename}-{work_all}-hours.xlsx"
        else:
            filename = f"devices-info-{start_date_filename}-to-{end_date_filename}-{work_all}-hours-{spaces_to_include}.xlsx"
        save_path = os.path.join(save_location, filename)
        df.to_excel(save_path, index=False)
        return save_path

    # Function to generate and format the report
    def generate_devices_report(save_location, start_date, end_date, work_hours_only, spaces_to_include, spaces):
        # If not specified, sets default end date to today and start date to month before end date. Turns date into
        # millisecond format. Makes date sting to be used in filename
        if len(end_date) == 0:
            end_date_string = datetime.today().strftime('%d/%m/%Y')
            end_date_datetime = datetime.strptime(end_date_string, "%d/%m/%Y")
        else:
            end_date_datetime = datetime.strptime(end_date, "%d/%m/%Y")
        end_date_mill = end_date_datetime.timestamp() * 1000
        if len(start_date) == 0:
            start_date_mill = (end_date_datetime.timestamp() - 24 * 60 * 60 * 30) * 1000
            start_date_datetime = datetime.fromtimestamp(start_date_mill / 1000)
        else:
            start_date_datetime = datetime.strptime(start_date, "%d/%m/%Y")
            start_date_mill = start_date_datetime.timestamp() * 1000

        try:
            all_devices_df = pd.DataFrame()

            num_spaces = len(spaces_df["id"])
            progress_bar["maximum"] = num_spaces

            with ThreadPoolExecutor(max_workers=2) as executor:
                future_to_space = {
                    executor.submit(get_devices, space_id, start_date_mill, end_date_mill, work_hours_only): space_id
                    for
                    space_id in
                    spaces["id"]}

                for future in as_completed(future_to_space):
                    space_id = future_to_space[future]
                    try:
                        devices_df = future.result()
                    except Exception as e:
                        print(f"Error occurred while getting devices for '{space_id}': {str(e)}")
                        traceback.print_exc()
                        continue

                    if not devices_df.empty:  # Only append if devices_df is not empty
                        all_devices_df = pd.concat([all_devices_df, devices_df], ignore_index=True)

                    progress_bar.step(1)
                    window.update()
            # Reorder the DataFrame columns
            all_devices_df = all_devices_df[
                [
                    'space_name',
                    'name',
                    'Link to Device',
                    'serialNumber',
                    'product_model',
                    'isOnline',
                    'sensorC',
                    'aveCO2',
                    'maxReadingCO2',
                    'dateOfMaxCO2',
                    'timeOfMaxCO2',
                    'minReadingCO2',
                    'dateOfMinCO2',
                    'timeOfMinCO2',
                    '%timeInHighRiskCO2',
                    'sensorH',
                    'aveHumidity',
                    'maxReadingHumidity',
                    'dateOfMaxHumidity',
                    'timeOfMaxHumidity',
                    'minReadingHumidity',
                    'dateOfMinHumidity',
                    'timeOfMinHumidity',
                    '%timeInHighRiskHumidity',
                    'sensorT',
                    'aveTemp',
                    'maxReadingTemp',
                    'dateOfMaxTemp',
                    'timeOfMaxTemp',
                    'minReadingTemp',
                    'dateOfMinTemp',
                    'timeOfMinTemp',
                    '%timeInHighRiskTemp',
                    'sensorACH',
                    'aveACH',
                    'maxReadingACH',
                    'dateOfMaxACH',
                    'timeOfMaxACH',
                    'minReadingACH',
                    'dateOfMinACH',
                    'timeOfMinACH',
                    'sensorVP',
                    'aveVP_Diff',
                    'maxReadingVP_Diff',
                    'dateOfMaxVP_Diff',
                    'timeOfMaxVP_Diff',
                    'minReadingVP_Diff',
                    'dateOfMinVP_Diff',
                    'timeOfMinVP_Diff',
                    '%timeInHighRiskVP_Diff',
                    'Fuel Poverty Risk',
                    'Void Risk'
                ]
            ]

            # Sort DataFrame by 'space_name'
            all_devices_df.sort_values('space_name', ascending=True, inplace=True)

            # Reset index after sorting
            all_devices_df.reset_index(drop=True, inplace=True)

            # Creates file
            save_path = create_excel_file(all_devices_df, save_location, start_date_datetime, end_date_datetime,
                                          work_hours_only, spaces_to_include)

            # Opens sheet to be edited
            wb = load_workbook(save_path)
            ws = wb['Sheet1']
            ws.insert_rows(idx=1)
            apply_header_style(ws)

            # Finds number of columns
            name_column_found = False
            link_column_found = False
            column_counter = 1
            empty_checker = ws.cell(row=2, column=1).value
            while empty_checker is not None or not link_column_found or not name_column_found:
                # Finds number of columns and index of device name and link column
                empty_checker = ws.cell(row=2, column=column_counter).value
                column_letter = get_column_letter(column_counter)
                if not name_column_found and empty_checker.startswith("name"):
                    name_column_found = True
                    name_column_letter = column_letter
                if not link_column_found and "Link" in empty_checker:
                    link_column_found = True
                    link_column_letter = column_letter
                    link_column_number = column_counter
                if empty_checker is not None and "aveCO2" in empty_checker:
                    beginning_of_data_letter = get_column_letter(column_counter - 1)
                column_counter = column_counter + 1
            column_counter = column_counter - 1  # One more than number of columns of data
            rows_devs = len(all_devices_df["name"]) + 2  # Number of rows
            for row in range(3, rows_devs + 1):
                # Adds link to device's homepage
                link = ws[f'{link_column_letter}{row}'].value
                name = ws[f'{name_column_letter}{row}'].value
                ws[f'{name_column_letter}{row}'].hyperlink = link
                ws[f'{name_column_letter}{row}'].value = name
                ws[f'{name_column_letter}{row}'].style = "Hyperlink"
            ws.delete_cols(link_column_number)
            column_counter = column_counter - 1
            last_column_letter = get_column_letter(column_counter - 1)

            # Creates borders, fonts, and backgrounds
            thin_border = Border(left=Side(style='thin', color="D9D9D9D9"),
                                 right=Side(style='thin', color="D9D9D9D9"),
                                 top=Side(style='thin', color="D9D9D9D9"),
                                 bottom=Side(style='thin', color="D9D9D9D9"))
            no_top_thin_border = Border(left=Side(style='thin', color="D9D9D9D9"),
                                        right=Side(style='thin', color="D9D9D9D9"),
                                        bottom=Side(style='thin', color="D9D9D9D9"))
            black_font = Font(color="FF000000", italic=False)
            black_italic_font = Font(color="FF000000", italic=True)
            white_font = Font(color="FFFFFFFF", italic=False)
            white_background = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type='solid')
            black_background = PatternFill(start_color="FF000000", end_color="FF000000", fill_type='solid')
            red_background = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type='solid')
            yellow_background = PatternFill(start_color="00FFC000", end_color="00FFC000", fill_type='solid')
            blue_background = PatternFill(start_color="0000B0F0", end_color="0000B0F0", fill_type='solid')
            purple_background = PatternFill(start_color="00D55EFF", end_color="00D55EFF")

            # Sets differential styles to be added to cells
            out_of_range_style = DifferentialStyle(fill=purple_background, font=black_italic_font)
            black_style = DifferentialStyle(fill=black_background, font=white_font)
            red_style = DifferentialStyle(fill=red_background, font=black_font)
            yellow_style = DifferentialStyle(fill=yellow_background, font=black_font)
            blue_style = DifferentialStyle(fill=blue_background, font=black_font)

            # Keeps dates, times, and N/As from being colored/italicized
            NA_style = DifferentialStyle(fill=white_background, border=thin_border, font=black_font)
            if rows_devs > 3:
                rule = Rule(type='containsText', operator='containsText', text='/', dxf=NA_style)
                rule.formula = [f'NOT(ISERROR(SEARCH("/",{beginning_of_data_letter}4)))']
                ws.conditional_formatting.add(f"{beginning_of_data_letter}4:{last_column_letter}{rows_devs}", rule)
                rule = Rule(type='containsText', operator='containsText', text=':', dxf=NA_style)
                rule.formula = [f'NOT(ISERROR(SEARCH(":",{beginning_of_data_letter}4)))']
                ws.conditional_formatting.add(f"{beginning_of_data_letter}4:{last_column_letter}{rows_devs}", rule)
            NA_style = DifferentialStyle(fill=white_background, border=no_top_thin_border, font=black_font)
            rule = Rule(type='containsText', operator='containsText', text='/', dxf=NA_style)
            rule.formula = [f'NOT(ISERROR(SEARCH("/",{beginning_of_data_letter}3)))']
            ws.conditional_formatting.add(f"{beginning_of_data_letter}3:{last_column_letter}3", rule)
            rule = Rule(type='containsText', operator='containsText', text=':', dxf=NA_style)
            rule.formula = [f'NOT(ISERROR(SEARCH(":",{beginning_of_data_letter}3)))']
            ws.conditional_formatting.add(f"{beginning_of_data_letter}3:{last_column_letter}3", rule)

            # Colors cells and renames column headers
            for col in range(1, column_counter):
                column_letter = get_column_letter(col)
                if "sensor" in ws.cell(row=2, column=col).value:
                    ws[f"{column_letter}2"] = "Parameter"
                    merge_start = column_letter
                elif "timeOfMinACH" in ws.cell(row=2, column=col).value:
                    # Merges cells for ACH header
                    merge_end = column_letter
                    ws.merge_cells(f'{merge_start}1:{merge_end}1')
                    ws[f"{merge_start}1"] = "ACH"
                elif "HighRisk" in ws.cell(row=2, column=col).value:
                    # Merges header cells and renames high risk cols and merged headers
                    merge_end = column_letter
                    ws.merge_cells(f'{merge_start}1:{merge_end}1')
                    if "CO2" in ws.cell(row=2, column=col).value:
                        ws[f"{merge_start}1"] = "CO2"
                    elif "Humidity" in ws.cell(row=2, column=col).value:
                        ws[f"{merge_start}1"] = "Humidity"
                    elif "Temp" in ws.cell(row=2, column=col).value:
                        ws[f"{merge_start}1"] = "Temperature"
                    elif "VP_Diff" in ws.cell(row=2, column=col).value:
                        ws[f"{merge_start}1"] = "VP Diff"
                    ws[f"{column_letter}2"] = "% Time in High Risk"
                elif "CO2" in ws.cell(row=2, column=col).value:  # Adds color to CO2 cells
                    rule = Rule(type='cellIs', operator='greaterThan', formula=[10000],
                                dxf=out_of_range_style)  # CO2 > 10,000
                    rule.formula = [f"{column_letter}3>10000"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=black_style)  # 10,001 > CO2 > 1500
                    rule.formula = [f"{column_letter}3>1500"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=red_style)  # 1501 > CO2 > 1000
                    rule.formula = [f"{column_letter}3>1000"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=yellow_style)  # 1001 > CO2 > 800
                    rule.formula = [f"{column_letter}3>800"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=out_of_range_style)  # 400 > CO2
                    rule.formula = [f"{column_letter}3<400"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                elif "Humidity" in ws.cell(row=2, column=col).value:  # Adds color to humidity cells
                    rule = Rule(type='cellIs', operator='greaterThan', formula=[99],
                                dxf=out_of_range_style)  # humidity > 99
                    rule.formula = [f"{column_letter}3>99"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=red_style)  # 100 > humidity > 60
                    rule.formula = [f"{column_letter}3>60"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=yellow_style)  # 40 > humidity > 5
                    rule.formula = [f"{column_letter}3<40"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=out_of_range_style)  # 5 > humidity
                    rule.formula = [f"{column_letter}3<5"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                elif "Temp" in ws.cell(row=2, column=col).value:  # Adds color to temperature cells
                    rule = Rule(type='cellIs', operator='greaterThan', formula=[40],
                                dxf=out_of_range_style)  # temperature > 40
                    rule.formula = [f"{column_letter}3>40"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=red_style)  # 41 > temperature > 25
                    rule.formula = [f"{column_letter}3>25"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=blue_style)  # 16 > temperature > -21
                    rule.formula = [f"{column_letter}3<16"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=out_of_range_style)  # -20 > temperature
                    rule.formula = [f"{column_letter}3<-20"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                elif "VP_Diff" in ws.cell(row=2, column=col).value:  # Adds color to VP diff cells
                    rule = Rule(type="expression", dxf=black_style)  # VP diff > 0.8
                    rule.formula = [f"{column_letter}3>0.8"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=red_style)  # 0.8 >= VP diff > 0.6
                    rule.formula = [f"{column_letter}3>0.6"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)
                    rule = Rule(type="expression", dxf=yellow_style)  # 0.6 >= VP diff > 0.3
                    rule.formula = [f"{column_letter}3>0.3"]
                    ws.conditional_formatting.add(f"{column_letter}3:{column_letter}{rows_devs}", rule)

                # Renames header cells within each parameter
                if "ave" in ws.cell(row=2, column=col).value:
                    ws[f"{column_letter}2"] = "Average"
                elif "maxReading" in ws.cell(row=2, column=col).value:
                    ws[f"{column_letter}2"] = "Max Reading"
                elif "dateOfMax" in ws.cell(row=2, column=col).value:
                    ws[f"{column_letter}2"] = "Date of Max"
                elif "timeOfMax" in ws.cell(row=2, column=col).value:
                    ws[f"{column_letter}2"] = "Time of Max"
                elif "minReading" in ws.cell(row=2, column=col).value:
                    ws[f"{column_letter}2"] = "Min Reading"
                elif "dateOfMin" in ws.cell(row=2, column=col).value:
                    ws[f"{column_letter}2"] = "Date of Min"
                elif "timeOfMin" in ws.cell(row=2, column=col).value:
                    ws[f"{column_letter}2"] = "Time of Min"

            ws.freeze_panes = 'A3'  # Freezes top two rows
            wb.save(save_path)  # Saves and closes spreadsheet

            print(f"Report generated successfully: {save_path}")

        except Exception as e:
            print(f"An error occurred: {str(e)}")
            traceback.print_exc()

    # "Generating report x/x" label
    report_track_label = tk.Label(window, text=report_track)
    report_track_label.grid(column=0, row=21, pady=3)

    # Progress bar
    progress_bar = ttk.Progressbar(window, orient=tk.HORIZONTAL, length=300, mode='determinate')
    progress_bar.grid(column=0, row=22, pady=3)

    generate_devices_report(save_location_entry, start_date_entry, end_date_entry, work_hours_only_var,
                            spaces_to_include_entry, spaces_df)

    # Removes label and progress bar
    report_track_label.grid_remove()
    progress_bar.grid_remove()
