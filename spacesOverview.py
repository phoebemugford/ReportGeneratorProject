import os
import pandas as pd
import requests
import json
from datetime import datetime
import time
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util import Retry
from concurrent.futures import ThreadPoolExecutor, as_completed
import traceback
import tkinter as tk
from tkinter import filedialog, ttk, messagebox


def spaces_overview_report(save_location_entry, spaces_to_include_entry, spaces_df, window, report_track):
    AUTH = "Authentication Key"
    URL = "API.com"
    HEADERS = {"Authorization": f"Bearer {AUTH}"}
    SESSION = requests.Session()
    RETRY = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
    SESSION.mount('http://', HTTPAdapter(max_retries=RETRY))
    SESSION.mount('https://', HTTPAdapter(max_retries=RETRY))

    # Gets and returns space info for each space
    def get_space_info(space_id):
        url = f"{URL}devices&space.id={space_id}&limit=2000"
        res = SESSION.get(url, headers=HEADERS)
        print(res.status_code)
        data_json = json.loads(res.content)
        devices = data_json["devices"]
        spaces_info = []
        online_count = 0
        offline_count = 0
        space_name = None
        has_weather = False
        weather_online = False
        # Goes through devices of each space
        for device in devices:
            product_model = device.get('product', {}).get('model') if device.get('product') else None
            space_name = device.get('space', {}).get('name')
            if product_model:
                if product_model.startswith('v1.0'):
                    has_weather = True
                    if device.get('isOnline'):
                        weather_online = True
                else:  # Counts number of online vs offline devices, ignoring weather stations
                    if device.get('isOnline'):
                        online_count = online_count + 1
                    else:
                        offline_count = offline_count + 1
            else:  # Still counts devices that don't have product model
                if device.get('isOnline'):
                    online_count = online_count + 1
                else:
                    offline_count = offline_count + 1
        # If there are no devices in the space, gets name from new API call
        if space_name is None:
            url = f"{URL}spaces/{space_id}"
            res = SESSION.get(url, headers=HEADERS)
            print(res.status_code)
            data_json = json.loads(res.content)
            space_name = data_json['name']
        link_to_space = f'dashboard.com/{space_id}'
        space_info = {'Name': space_name, 'Link to Space': link_to_space, 'Devices Online': online_count,
                      'Devices Offline': offline_count, 'Weather Station': has_weather, 'Weather Online': weather_online}
        spaces_info.append(space_info)
        time.sleep(1)
        spaces_info_df = pd.DataFrame(spaces_info)
        return spaces_info_df

    def apply_header_style(ws):
        # Define style for header
        header = NamedStyle(name="header")
        header.font = Font(bold=True)
        header.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                               top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        header.alignment = Alignment(horizontal="center")

        # Apply style to headers
        for cell in ws[1]:
            cell.style = header

    # Function to create the Excel file
    def create_excel_file(df, save_location, spaces_to_include):
        # Sets variables for filename
        today = datetime.today().strftime('%d-%m-%Y')
        if len(spaces_to_include) == 0:
            filename = f"spaces-overview-{today}.xlsx"
        else:
            filename = f"spaces-overview-{today}-{spaces_to_include}.xlsx"
        save_path = os.path.join(save_location, filename)
        df.to_excel(save_path, index=False)
        return save_path

    # Function to generate and format the report
    def generate_spaces_report(save_location, spaces_to_include, spaces):
        try:
            all_spaces_df = pd.DataFrame()

            num_spaces = len(spaces_df["id"])
            progress_bar["maximum"] = num_spaces

            with ThreadPoolExecutor(max_workers=2) as executor:
                future_to_space = {executor.submit(get_space_info, space_id): space_id for space_id in
                                   spaces["id"]}

                for future in as_completed(future_to_space):
                    space_id = future_to_space[future]
                    try:
                        space_infos_df = future.result()
                    except Exception as e:
                        print(f"Error occurred while getting info for '{space_id}': {str(e)}")
                        traceback.print_exc()
                        continue

                    if not space_infos_df.empty:  # Only append if space_infos_df is not empty
                        all_spaces_df = pd.concat([all_spaces_df, space_infos_df], ignore_index=True)

                    progress_bar.step(1)
                    window.update()

            # Then reorder the DataFrame columns
            all_spaces_df = all_spaces_df[
                [
                    'Name',
                    'Link to Space',
                    'Devices Online',
                    'Devices Offline',
                    'Weather Station',
                    'Weather Online'
                ]
            ]

            # Sort DataFrame by 'space_name'
            all_spaces_df.sort_values('Name', ascending=True, inplace=True)

            # Reset index after sorting
            all_spaces_df.reset_index(drop=True, inplace=True)

            save_path = create_excel_file(all_spaces_df, save_location, spaces_to_include)

            # Adds link to space name
            wb = load_workbook(save_path)
            ws = wb['Sheet1']
            column_finder = 1
            name_column_found = False
            link_column_found = False
            while not link_column_found or not name_column_found:
                column_letter = get_column_letter(column_finder)
                if "Name" in ws[f'{column_letter}1'].value:
                    name_column_found = True
                    name_column_letter = column_letter
                if "Link" in ws[f'{column_letter}1'].value:
                    link_column_found = True
                    link_column_letter = column_letter
                    link_column_number = column_finder
                column_finder = column_finder + 1
            rows_devs = len(all_spaces_df["Name"]) + 1
            for row in range(2, rows_devs+1):
                link = ws[f'{link_column_letter}{row}'].value
                name = ws[f'{name_column_letter}{row}'].value
                ws[f'{name_column_letter}{row}'].hyperlink = link
                ws[f'{name_column_letter}{row}'].value = name
                ws[f'{name_column_letter}{row}'].style = "Hyperlink"
            ws.delete_cols(link_column_number)
            ws.freeze_panes = 'A2'  # Freezes top row
            wb.save(save_path)

            print(f"Report generated successfully: {save_path}")

        except Exception as e:
            print(f"An error occurred: {str(e)}")
            traceback.print_exc()

    # "Generating report x/x" label
    report_track_label = tk.Label(window, text=report_track)
    report_track_label.grid(column=0, row=19, pady=3)

    # Progress bar
    progress_bar = ttk.Progressbar(window, orient=tk.HORIZONTAL, length=300, mode='determinate')
    progress_bar.grid(column=0, row=20, pady=3)

    time.sleep(2)

    generate_spaces_report(save_location_entry, spaces_to_include_entry, spaces_df)

    # Removes label and progress bar
    report_track_label.grid_remove()
    progress_bar.grid_remove()
